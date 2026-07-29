#!/usr/bin/env python3
from __future__ import annotations
import json, sqlite3, pathlib, re, hashlib
from datetime import datetime
from openpyxl import load_workbook

ROOT=pathlib.Path(__file__).resolve().parents[1]
REPORTS=ROOT/'reports'
SCHEMA_FILES=['schema.sql','schema2.sql','schema4.sql','schema7.sql','alter_users.sql','data.sql','migrations/0008_connected_erp.sql','migrations/0010_procurement_sales_controls.sql','migrations/0011_finance_planning_registers.sql']

def scalar(db,sql,args=()): return db.execute(sql,args).fetchone()[0]

def main():
    db=sqlite3.connect(':memory:');db.row_factory=sqlite3.Row
    tests=[]
    def test(name,condition,detail=''):
        tests.append((name,bool(condition),str(detail)))
        if not condition: print('FAIL',name,detail)
    for f in SCHEMA_FILES:
        db.executescript((ROOT/f).read_text(encoding='utf-8',errors='ignore'))
    chunks=sorted((ROOT/'migrations/opening').glob('*_opening_data.sql'))
    for f in chunks: db.executescript(f.read_text(encoding='utf-8'))
    db.execute('PRAGMA foreign_keys=ON')

    fk=db.execute('PRAGMA foreign_key_check').fetchall()
    test('All opening SQL chunks execute',len(chunks)>0,f'{len(chunks)} chunks')
    test('No foreign-key violations',len(fk)==0,f'{len(fk)} violations')
    test('Actual source workbooks embedded',scalar(db,'select count(*) from erp_opening_data_control')==14,scalar(db,'select count(*) from erp_opening_data_control'))
    test('Source rows archived',scalar(db,'select count(*) from erp_import_rows')>=20000,scalar(db,'select count(*) from erp_import_rows'))
    test('Canonical assets loaded',scalar(db,'select count(*) from erp_assets')>=8000,scalar(db,'select count(*) from erp_assets'))
    test('No duplicate canonical asset serials',scalar(db,'select count(*) from (select serial_no from erp_assets group by serial_no having count(*)>1)')==0)
    test('Duplicate serial evidence preserved',scalar(db,"select count(*) from erp_serial_exceptions where exception_type='DUPLICATE_MASTER_SERIAL'")>0,scalar(db,'select count(*) from erp_serial_exceptions'))
    test('Shipments created from STELLAR/ATLAS',scalar(db,'select count(*) from erp_shipments')>=20,scalar(db,'select count(*) from erp_shipments'))
    test('ATLAS expected serials created',scalar(db,'select count(*) from erp_expected_assets')>=1900,scalar(db,'select count(*) from erp_expected_assets'))
    test('Historical receiving linked to shipments',scalar(db,'select count(*) from erp_receipt_lines')>=1900,scalar(db,'select count(*) from erp_receipt_lines'))
    test('Stock movements loaded',scalar(db,'select count(*) from erp_stock_ledger')>=5000,scalar(db,'select count(*) from erp_stock_ledger'))
    test('Sales and lease assignments loaded',scalar(db,'select count(*) from erp_sales_orders')>=250,scalar(db,'select count(*) from erp_sales_orders'))
    test('Delivery asset history loaded',scalar(db,'select count(*) from erp_delivery_assets')>=1000,scalar(db,'select count(*) from erp_delivery_assets'))
    test('Historical returns loaded',scalar(db,'select count(*) from erp_return_orders')>0,scalar(db,'select count(*) from erp_return_orders'))
    test('Battery swaps remain unreconciled',scalar(db,"select count(*) from erp_reconciliation_cases where status='UNRECONCILED'")>0,scalar(db,"select count(*) from erp_reconciliation_cases where status='UNRECONCILED'"))
    test('Requisitions loaded',scalar(db,'select count(*) from erp_requisitions')>=200,scalar(db,'select count(*) from erp_requisitions'))
    test('Pre-release checks loaded',scalar(db,'select count(*) from erp_pre_release_checks')>=250,scalar(db,'select count(*) from erp_pre_release_checks'))
    test('Approved purchase orders loaded',scalar(db,'select count(*) from erp_purchase_orders')>=15,scalar(db,'select count(*) from erp_purchase_orders'))
    test('Procurement register loaded',scalar(db,'select count(*) from erp_procurement_register')>=1200,scalar(db,'select count(*) from erp_procurement_register'))
    test('Sales receipts loaded',scalar(db,'select count(*) from erp_sales_receipts')>=100,scalar(db,'select count(*) from erp_sales_receipts'))
    test('Budget loaded',scalar(db,'select count(*) from erp_budget_plan')>0,scalar(db,'select count(*) from erp_budget_plan'))
    test('Planning drivers loaded',scalar(db,'select count(*) from erp_planning_drivers')>=4000,scalar(db,'select count(*) from erp_planning_drivers'))
    test('Station projects linked to assets',scalar(db,'select count(*) from erp_station_project_assets')>=300,scalar(db,'select count(*) from erp_station_project_assets'))
    test('Admin live access loaded',scalar(db,"select count(*) from erp_users where email='mmungcal@nrdev.ph' and role_code='ADMIN' and live_access=1")==1)

    # Every serialized transaction must point to a canonical asset when the serial exists in the opening data.
    orphan_sales=scalar(db,"select count(*) from erp_sales_lines l where coalesce(l.serial_no,'')<>'' and l.asset_id is null")
    orphan_delivery=scalar(db,"select count(*) from erp_delivery_assets d where coalesce(d.serial_no,'')<>'' and d.asset_id is null")
    test('Sales serials resolve to assets',orphan_sales==0,orphan_sales)
    test('Delivery serials resolve to assets',orphan_delivery==0,orphan_delivery)

    # Item auto-code sequence safety.
    sequence_ok=True;details=[]
    for cat in ['MC','BAT','BSS','SP','CHG','OTH']:
        mx=0
        for (code,) in db.execute('select item_code from erp_items where item_code like ?',(cat+'-%',)):
            m=re.fullmatch(re.escape(cat)+r'-(\d+)',code or '')
            if m: mx=max(mx,int(m.group(1)))
        nxt=scalar(db,'select next_value from erp_sequences where code=?',('ITEM_'+cat,))
        details.append(f'{cat}:max={mx},next={nxt}')
        sequence_ok &= nxt>mx
    test('New item codes cannot collide with opening codes',sequence_ok,'; '.join(details))

    # Password values must not be copied to the source archive.
    leaked=scalar(db,"select count(*) from erp_import_rows where payload_json like '%Password1%' or payload_json like '%Password2%' or payload_json like '%Hello123%'")
    test('Legacy plaintext passwords redacted from database archive',leaked==0,leaked)

    # The GitHub/deployment copy of the checklist must retain operations data but not readable credentials.
    checklist=ROOT/'source_data'/'Pre-release Unit Checklist.xlsx'
    source_password_leaks=[]
    wb=load_workbook(checklist,read_only=True,data_only=False)
    try:
        for ws in wb.worksheets:
            password_columns=[]
            for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row or 0,20)):
                for cell in row:
                    if isinstance(cell.value,str) and cell.value.strip().lower() in {'password','passwd'}:
                        password_columns.append((cell.column,cell.row))
            for col,header_row in password_columns:
                for row_no in range(header_row+1,(ws.max_row or header_row)+1):
                    value=ws.cell(row_no,col).value
                    if value not in (None,'','[REDACTED]'):
                        source_password_leaks.append(f'{ws.title}!{ws.cell(row_no,col).coordinate}')
    finally:
        wb.close()
    test('Bundled source workbooks contain no readable passwords',len(source_password_leaks)==0,len(source_password_leaks))

    counts={}
    for t in ['erp_items','erp_assets','erp_serial_exceptions','erp_shipments','erp_expected_assets','erp_receipts','erp_receipt_lines','erp_stock_ledger','erp_sales_orders','erp_sales_lines','erp_deliveries','erp_delivery_assets','erp_return_orders','erp_return_lines','erp_reconciliation_cases','erp_requisitions','erp_requisition_lines','erp_pre_release_checks','erp_purchase_orders','erp_landed_cost_headers','erp_station_projects','erp_station_project_assets','erp_sales_receipts','erp_procurement_register','erp_payment_register','erp_budget_plan','erp_planning_drivers','erp_import_rows']:
        counts[t]=scalar(db,f'select count(*) from {t}')
    inventory_status={r['current_status']:r['n'] for r in db.execute('select current_status,count(*) n from erp_assets group by current_status order by n desc')}
    categories={r['category']:r['n'] for r in db.execute('select category,count(*) n from erp_assets group by category order by n desc')}
    source_rows={r['source_name']:r['source_rows'] for r in db.execute('select source_name,source_rows from erp_opening_data_control order by source_name')}

    passed=sum(1 for _,ok,_ in tests if ok); total=len(tests)
    REPORTS.mkdir(exist_ok=True)
    lines=['# E88 FinSys v7.1 Self-Test Report','',f'Generated: {datetime.now().isoformat(timespec="seconds")}',f'**Result: {passed}/{total} tests passed.**','']
    for name,ok,detail in tests:
        lines.append(f'- [{"x" if ok else " "}] **{name}** — {detail}')
    lines += ['','## Loaded operational counts','', '| Table | Rows |','|---|---:|']+[f'| `{k}` | {v:,} |' for k,v in counts.items()]
    lines += ['','## Inventory by category','', '| Category | Assets |','|---|---:|']+[f'| {k} | {v:,} |' for k,v in categories.items()]
    lines += ['','## Inventory by current status','', '| Status | Assets |','|---|---:|']+[f'| {k} | {v:,} |' for k,v in inventory_status.items()]
    lines += ['','## Test boundary','', 'The package was tested locally against SQLite using the complete schema, legacy opening data, connected ERP migrations, and all generated opening-data chunks. Live Cloudflare Access, R2 uploads, D1 concurrency, and the production Workers deployment must still be smoke-tested after deployment because those services are not available in the local container.']
    (REPORTS/'SELF_TEST_REPORT.md').write_text('\n'.join(lines)+'\n',encoding='utf-8')

    dl=['# E88 FinSys v7.1 Data Load Report','',f'Generated: {datetime.now().isoformat(timespec="seconds")}', '', '## Actual source workbooks', '', '| Source | Archived operational rows |','|---|---:|']
    dl += [f'| {k} | {v:,} |' for k,v in source_rows.items()]
    dl += ['','## Canonicalization policy','', '- One canonical asset is retained for each normalized serial number.', '- Duplicate master occurrences are preserved as open serial exceptions; they are not deleted.', '- Operational references across STAR, STAKU, SATURN, ATLAS, requisitions, checklists, and warehouse documents link back to the canonical asset.', '- Battery serial swaps on return are accepted into quarantine and remain `UNRECONCILED` until reviewed.', '- Missing item descriptions automatically receive category-based item codes. Runtime sequences are advanced past every migrated code.', '- Legacy password columns are redacted from both the database source archive and the deployable workbook copy. All non-credential operational fields remain included.']
    (REPORTS/'DATA_LOAD_REPORT.md').write_text('\n'.join(dl)+'\n',encoding='utf-8')
    (REPORTS/'SELF_TEST_REPORT.json').write_text(json.dumps({'passed':passed,'total':total,'tests':[{'name':n,'passed':o,'detail':d} for n,o,d in tests],'counts':counts,'categories':categories,'statuses':inventory_status},indent=2),encoding='utf-8')
    print(json.dumps({'passed':passed,'total':total,'failures':[n for n,o,d in tests if not o],'counts':counts},indent=2))
    if passed!=total: raise SystemExit(1)

if __name__=='__main__': main()
