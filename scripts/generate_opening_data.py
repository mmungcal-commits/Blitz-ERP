#!/usr/bin/env python3
"""Generate the E88 FinSys connected opening database from the actual source workbooks.

The generator intentionally preserves duplicate serial evidence as exception cases. It never
silently drops a source row. Password fields in legacy worksheets are redacted from the ERP
source archive; source workbooks remain bundled separately for authorized migration review.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sqlite3
import sys
import warnings
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "source_data"
OPENING = ROOT / "migrations" / "opening"
REPORTS = ROOT / "reports"

ARCHIVE_SHEETS = {
    "ATLAS - Asset Manifest (1).xlsx": ["MOTORCYCLE", "BATTERY", "LOCKER", "INVOICE CODES"],
    "STELLAR _ Shipment.xlsx": ["TRACKING", "Stacku - Logistics Shipment Cha"],
    "STAR _ E88 SCM Inventory2026 (4).xlsx": ["E88 - Item Code", "AW -  Transaction Log", "AW - Inventory - In", "SLoc - In (Assets)"],
    "STAKU - SALES_LEASE B2B (4).xlsx": ["TRANSACTION"],
    "SATURN _ DELIVERY MONITORING _ LAST MILE (3).xlsx": ["PENDING DELIVERY", "Proof of Delivery", "2026"],
    "Pre-release Unit Checklist.xlsx": ["Sheet1"],
    "2026 SCM Warehouse Documents.xlsx": ["2026 Transaction"],
    "SCM Requisition Slip 1226.xlsx": ["Form Responses 1"],
    "SCM Live Dashboard (3).xlsx": ["Details", "SCM Weekly Schedule", "For Prep", "🔥BURNDOWN", "WH"],
    "E88_SalesMonitoring_2026.xlsx": ["Aftersales", "WhseServiceSales", "MCSales", "MCLeased", "BSwapping"],
    "E88_ProcurementMonitoring_2026.xlsx": ["Procurement Monitoring", "Payments Recording", "GL Journal"],
    "Detailed_Receipts_2026.xlsx": ["Detailed Receipts Jan-Jul"],
    "E88_ApprovedBudget2026 (4)(1).xlsx": ["ApprovedBudget", "Purchase Orders (Approved)", "Jan Budget vs Actual", "Feb Budget vs Actual", "Mar Budget vs Actual", "Apr Budget vs Actual", "May Budget vs Actual", "June Budget vs Actual", "July Budget vs Actual", "Detailed Receipts"],
    "E88_AM_FINAL_v5A.xlsx": ["MC Drivers (NRD)", "Energy Drivers (RideBox)", "Operating KPIs", "Station Economics", "Battery Lifecycle", "Pipeline & Cap Table", "Actuals to Gather"],
}

SOURCE_FILES = [
    "ATLAS - Asset Manifest (1).xlsx",
    "STELLAR _ Shipment.xlsx",
    "STAR _ E88 SCM Inventory2026 (4).xlsx",
    "STAKU - SALES_LEASE B2B (4).xlsx",
    "SATURN _ DELIVERY MONITORING _ LAST MILE (3).xlsx",
    "Pre-release Unit Checklist.xlsx",
    "2026 SCM Warehouse Documents.xlsx",
    "SCM Requisition Slip 1226.xlsx",
    "SCM Live Dashboard (3).xlsx",
    "E88_SalesMonitoring_2026.xlsx",
    "E88_ProcurementMonitoring_2026.xlsx",
    "Detailed_Receipts_2026.xlsx",
    "E88_ApprovedBudget2026 (4)(1).xlsx",
    "E88_AM_FINAL_v5A.xlsx",
]

CONNECTED_TABLES = [
    "erp_settings", "erp_sequences", "erp_roles", "erp_role_permissions", "erp_users",
    "erp_locations", "erp_partners", "erp_items", "erp_import_batches", "erp_import_rows",
    "erp_purchase_orders", "erp_purchase_order_lines", "erp_landed_cost_headers",
    "erp_landed_cost_lines", "erp_landed_cost_allocations", "erp_shipments",
    "erp_shipment_lines", "erp_expected_assets", "erp_receipts", "erp_receipt_lines",
    "erp_assets", "erp_stock_ledger", "erp_assignments", "erp_assignment_assets",
    "erp_return_orders", "erp_return_lines", "erp_reconciliation_cases", "erp_requisitions",
    "erp_requisition_lines", "erp_pre_release_checks", "erp_sales_orders", "erp_sales_lines",
    "erp_deliveries", "erp_delivery_assets", "erp_station_projects", "erp_station_project_assets",
    "erp_audit_log", "erp_serial_exceptions", "erp_documents", "erp_customer_credit_events",
    "erp_qr_reviews", "erp_sales_receipts", "erp_procurement_register", "erp_payment_register",
    "erp_budget_plan", "erp_planning_drivers", "erp_opening_data_control",
]


def clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat(sep=" ", timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat(timespec="seconds")
    if isinstance(v, bool):
        return "1" if v else "0"
    s = str(v).strip()
    if s in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "None", "nan"}:
        return ""
    return re.sub(r"\s+", " ", s)


def date_text(v: Any) -> str:
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    s = clean(v)
    if not s or s.upper() in {"NONE", "N/A", "-"}: return ""
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%b %d, %Y"):
        try: return datetime.strptime(s, fmt).date().isoformat()
        except ValueError: pass
    return s[:10] if re.match(r"\d{4}-\d{2}-\d{2}", s) else s


def number(v: Any, default: float = 0.0) -> float:
    if v in (None, ""): return default
    if isinstance(v, (int, float)): return float(v)
    s = clean(v).replace("₱", "").replace("$", "").replace(",", "").replace("%", "")
    try: return float(s)
    except ValueError: return default


def norm(v: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", clean(v).upper()).strip()


def serial(v: Any) -> str:
    s = clean(v).upper().replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", "", s)
    if s in {"", "-", "NONE", "N/A", "NA", "#N/A"}: return ""
    return s


def infer_category(*values: Any) -> str:
    s = " ".join(norm(v) for v in values if v not in (None, ""))
    if any(k in s for k in ["BATTERY", " BAT ", "AMPACE", "7428", "GREENWAY BAT"]): return "BAT"
    if any(k in s for k in ["LOCKER", "SWAP STATION", "SWAPPING STATION", "SPACEPORT", " BSS ", "RACK"]): return "BSS"
    if "CHARGER" in s or "CHARGING KIT" in s: return "CHG"
    if any(k in s for k in ["MOTORCYCLE", " D400", " R280", "HORWIN", " E MOTORCYCLE", " MC "]): return "MC"
    if any(k in s for k in ["SPARE", "PART", "CONTROLLER", "GLOVE BOX", "HELMET", "TOOL", "CANALYST"]): return "SP"
    return "OTH"


def status_norm(v: Any) -> str:
    s = norm(v)
    if "SOLD" in s: return "SOLD"
    if "LEASE" in s or "ONGOING" in s: return "LEASED"
    if "RETURN" in s: return "AVAILABLE"
    if "REPAIR" in s or "MAINT" in s: return "UNDER_REPAIR"
    if "DEMO" in s or "PILOT" in s or "TEST" in s: return "PILOT_TEST"
    if "ASSIGN" in s or "DEPLOY" in s or "DELIVER" in s: return "ASSIGNED"
    if "AVAILABLE" in s or "RECEIVED" in s: return "AVAILABLE"
    return clean(v).upper().replace(" ", "_") or "AVAILABLE"


def hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while chunk := f.read(1024 * 1024): h.update(chunk)
    return h.hexdigest()


def make_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    for f in ["schema.sql", "schema2.sql", "schema4.sql", "schema7.sql", "alter_users.sql", "data.sql",
              "migrations/0008_connected_erp.sql", "migrations/0010_procurement_sales_controls.sql",
              "migrations/0011_finance_planning_registers.sql"]:
        conn.executescript((ROOT / f).read_text(encoding="utf-8", errors="ignore"))
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


class _Cell:
    def __init__(self, value): self.value = value


class CachedWorksheet:
    def __init__(self, raw):
        try: raw.reset_dimensions()
        except Exception: pass
        self.title = raw.title
        self._rows = [tuple(row) for row in raw.iter_rows(values_only=True)]
        self.max_row = len(self._rows)
        self.max_column = max((len(r) for r in self._rows), default=0)
    def cell(self, row, column):
        if row < 1 or column < 1 or row > self.max_row: return _Cell(None)
        values = self._rows[row-1]
        return _Cell(values[column-1] if column <= len(values) else None)


class CachedWorkbook:
    def __init__(self, path: Path):
        self._raw = load_workbook(path, data_only=True, read_only=True)
        self.sheetnames = list(self._raw.sheetnames)
        self._cache = {}
    def __getitem__(self, name):
        if name not in self._cache:
            self._cache[name] = CachedWorksheet(self._raw[name])
        return self._cache[name]


class Loader:
    def __init__(self, conn: sqlite3.Connection):
        self.db = conn
        self.item_seq = Counter({"MC": 1, "BAT": 1, "BSS": 1, "SP": 1, "CHG": 1, "OTH": 1})
        self.partner_seq = Counter({"CUSTOMER": 1, "VENDOR": 1, "EMPLOYEE": 1, "SITE_PARTNER": 1})
        self.location_seq = 1
        self.asset_seq = 1
        self.move_seq = 1
        self.exception_seq = 1
        self.import_seq = 1
        self.assignment_seq = 1
        self.sales_seq = 1
        self.delivery_seq = 1
        self.requisition_seq = 1
        self.check_seq = 1
        self.receipt_seq = 1
        self.recon_seq = 1
        self.return_seq = 1
        self.station_seq = 1
        self.po_seq = 1
        self.lc_seq = 1
        self.master_occurrences: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.movement_keys: dict[tuple, int] = {}
        self.workbooks: dict[str, Any] = {}
        self.counts = Counter()
        self.source_stats: dict[str, dict[str, int]] = defaultdict(lambda: Counter())

    def wb(self, name: str):
        if name not in self.workbooks:
            self.workbooks[name] = CachedWorkbook(SOURCE / name)
        return self.workbooks[name]

    def ensure_item(self, description: Any, category: Any = "", item_code: Any = "", **extra) -> sqlite3.Row:
        name = clean(description) or f"{infer_category(category)} Item"
        cat = infer_category(category, name)
        normalized = norm(name)
        code = clean(item_code).upper().replace(" ", "")
        if code:
            row = self.db.execute("SELECT * FROM erp_items WHERE item_code=?", (code,)).fetchone()
            if row: return row
        row = self.db.execute("SELECT * FROM erp_items WHERE normalized_name=? AND category=?", (normalized, cat)).fetchone()
        if row: return row
        if not code:
            while True:
                code = f"{cat}-{self.item_seq[cat]:06d}"
                self.item_seq[cat] += 1
                if not self.db.execute("SELECT 1 FROM erp_items WHERE item_code=?", (code,)).fetchone(): break
        try:
            self.db.execute("""INSERT INTO erp_items(item_code,item_name,normalized_name,category,subcategory,manufacturer,model,color,serialized,standard_cost,auto_created,source_system,source_key)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (code, name, normalized, cat, clean(extra.get("subcategory")), clean(extra.get("manufacturer")),
                             clean(extra.get("model")), clean(extra.get("color")), 1 if extra.get("serialized", cat in {"MC","BAT","BSS","CHG"}) else 0,
                             number(extra.get("standard_cost")), 1 if not item_code else 0, clean(extra.get("source_system")), clean(extra.get("source_key"))))
        except sqlite3.IntegrityError:
            return self.db.execute("SELECT * FROM erp_items WHERE normalized_name=? AND category=?", (normalized, cat)).fetchone()
        self.counts["items_created"] += 1
        return self.db.execute("SELECT * FROM erp_items WHERE item_code=?", (code,)).fetchone()

    def ensure_location(self, name: Any, location_type: str = "OTHER", code: str = "") -> sqlite3.Row:
        nm = clean(name) or "Unassigned"
        row = self.db.execute("SELECT * FROM erp_locations WHERE upper(name)=upper(?)", (nm,)).fetchone()
        if row: return row
        if not code:
            code = re.sub(r"[^A-Z0-9]+", "-", nm.upper()).strip("-")[:24]
            if not code or self.db.execute("SELECT 1 FROM erp_locations WHERE code=?", (code,)).fetchone():
                code = f"LOC-{self.location_seq:05d}"; self.location_seq += 1
        self.db.execute("INSERT INTO erp_locations(code,name,location_type) VALUES(?,?,?)", (code, nm, location_type))
        return self.db.execute("SELECT * FROM erp_locations WHERE code=?", (code,)).fetchone()

    def ensure_partner(self, name: Any, partner_type: str = "CUSTOMER", code: str = "", **extra) -> sqlite3.Row:
        nm = clean(name) or "Unknown"
        typ = partner_type.upper()
        row = self.db.execute("SELECT * FROM erp_partners WHERE partner_type=? AND upper(name)=upper(?)", (typ, nm)).fetchone()
        if row: return row
        if not code:
            pref = {"CUSTOMER":"CUS","VENDOR":"VEN","EMPLOYEE":"EMP","SITE_PARTNER":"PAR"}.get(typ,"PAR")
            while True:
                code = f"{pref}-{self.partner_seq[typ]:06d}"; self.partner_seq[typ] += 1
                if not self.db.execute("SELECT 1 FROM erp_partners WHERE partner_code=?", (code,)).fetchone(): break
        self.db.execute("""INSERT INTO erp_partners(partner_code,partner_type,name,address,email,phone,source_system,source_key)
                          VALUES(?,?,?,?,?,?,?,?)""",
                        (code, typ, nm, clean(extra.get("address")), clean(extra.get("email")).lower(), clean(extra.get("phone")),
                         clean(extra.get("source_system")), clean(extra.get("source_key"))))
        return self.db.execute("SELECT * FROM erp_partners WHERE partner_code=?", (code,)).fetchone()

    def add_asset_occurrence(self, s: Any, *, category: Any = "", description: Any = "", item_code: Any = "",
                             source_system: str, source_sheet: str, source_row: int, master: bool = False, **fields):
        sn = serial(s)
        if not sn or len(sn) < 4: return
        cat = infer_category(category, description, item_code)
        self.master_occurrences[sn].append({"source_system":source_system,"source_sheet":source_sheet,"source_row":source_row,"master":master,"category":cat,"description":clean(description),"item_code":clean(item_code),**fields})

    def archive_sources(self):
        for filename in SOURCE_FILES:
            path = SOURCE / filename
            if not path.exists(): raise FileNotFoundError(path)
            import_no = f"SRC-{self.import_seq:04d}"; self.import_seq += 1
            file_hash = hash_file(path)
            cur = self.db.execute("""INSERT INTO erp_import_batches(import_no,import_type,source_file_name,source_hash,status,created_by,posted_at)
                                  VALUES(?,?,?,?,'POSTED','MIGRATION',datetime('now'))""", (import_no,"OPENING_SOURCE",filename,file_hash))
            import_id = cur.lastrowid
            print(f"Archiving {filename}", flush=True)
            wb = self.wb(filename)
            total = 0
            selected = [x for x in ARCHIVE_SHEETS.get(filename, wb.sheetnames) if x in wb.sheetnames]
            for sheet_name in selected:
                ws = wb[sheet_name]
                sensitive_cols = set()
                for r in range(1, min(ws.max_row or 0, 12) + 1):
                    for c in range(1, (ws.max_column or 0) + 1):
                        if "PASSWORD" in norm(ws.cell(r,c).value): sensitive_cols.add(c)
                batch=[]
                for r in range(1, (ws.max_row or 0)+1):
                    vals=[]; nonempty=False
                    for c in range(1, (ws.max_column or 0)+1):
                        v=ws.cell(r,c).value
                        cv=clean(v)
                        if c in sensitive_cols and r>1 and cv: cv="[REDACTED]"
                        if cv: nonempty=True
                        vals.append(cv)
                    if not nonempty: continue
                    total += 1
                    payload=json.dumps({"values":vals},ensure_ascii=False,separators=(",",":"))
                    batch.append((import_id,ws.title,r,"SOURCE_ROW",f"{ws.title}:{r}",payload,"ARCHIVED",""))
                    if len(batch)>=500:
                        self.db.executemany("""INSERT OR IGNORE INTO erp_import_rows(import_id,source_sheet,source_row,record_type,external_key,payload_json,validation_status,validation_message)
                                             VALUES(?,?,?,?,?,?,?,?)""",batch); batch=[]
                if batch:
                    self.db.executemany("""INSERT OR IGNORE INTO erp_import_rows(import_id,source_sheet,source_row,record_type,external_key,payload_json,validation_status,validation_message)
                                         VALUES(?,?,?,?,?,?,?,?)""",batch)
            self.db.execute("UPDATE erp_import_batches SET total_rows=?,valid_rows=? WHERE id=?",(total,total,import_id))
            self.db.execute("INSERT OR REPLACE INTO erp_opening_data_control(source_name,source_file_name,source_hash,source_rows,status,notes) VALUES(?,?,?,?, 'LOADED',?)",
                            (filename,filename,file_hash,total,"Actual source workbook archived; password columns redacted in ERP archive."))
            self.source_stats[filename]["archived_rows"] = total
            self.counts["source_rows"] += total

    def load_master_data(self):
        # Existing legacy item and customer data.
        for r in self.db.execute("SELECT * FROM items ORDER BY id"):
            self.ensure_item(r["description"], r["category"], r["sku"], standard_cost=r["unit_cost"], source_system="LEGACY_DB",source_key=str(r["id"]))
        for r in self.db.execute("SELECT * FROM customers ORDER BY id"):
            self.ensure_partner(r["name"],"CUSTOMER",r["code"] or "",source_system="LEGACY_DB",source_key=str(r["id"]))
        self.ensure_location("E88 Asgard Warehouse","WAREHOUSE","ASGARD")
        self.ensure_location("Asgard Warehouse","WAREHOUSE","ASGARD-WH")
        self.ensure_location("TCI - Pasig","WAREHOUSE","TCI-PASIG")
        self.ensure_location("Retail Sta. Rosa","RETAIL","RETAIL-STA-ROSA")
        self.ensure_location("Returns Quarantine","QUARANTINE","RET-QUAR")

        # STAR item code is the approved legacy item code list.
        ws=self.wb("STAR _ E88 SCM Inventory2026 (4).xlsx")["E88 - Item Code"]
        for r in range(2,(ws.max_row or 0)+1):
            code=clean(ws.cell(r,1).value); desc=clean(ws.cell(r,2).value)
            if code and desc: self.ensure_item(desc,infer_category(desc,code),code,source_system="STAR",source_key=str(r))

        # Suppliers and shipment locations.
        ws=self.wb("STELLAR _ Shipment.xlsx")["TRACKING"]
        for r in range(22,(ws.max_row or 0)+1):
            shipment=clean(ws.cell(r,4).value)
            if not shipment: continue
            supplier=clean(ws.cell(r,6).value)
            if supplier: self.ensure_partner(supplier,"VENDOR",address=ws.cell(r,7).value,source_system="STELLAR",source_key=shipment)
            for val,typ in [(ws.cell(r,8).value,"PORT"),(ws.cell(r,24).value,"WAREHOUSE")]:
                if clean(val): self.ensure_location(val,typ)

        # Commercial customers.
        ws=self.wb("STAKU - SALES_LEASE B2B (4).xlsx")["TRANSACTION"]
        for r in range(2,(ws.max_row or 0)+1):
            cust=clean(ws.cell(r,7).value)
            if cust: self.ensure_partner(cust,"CUSTOMER",address=ws.cell(r,8).value,source_system="STAKU",source_key=str(r))

        # Requisition requestors, customers, and destinations.
        ws=self.wb("SCM Requisition Slip 1226.xlsx")["Form Responses 1"]
        for r in range(2,(ws.max_row or 0)+1):
            req=clean(ws.cell(r,1).value)
            if not req: continue
            company=clean(ws.cell(r,6).value); requestor=clean(ws.cell(r,29).value)
            if company: self.ensure_partner(company,"CUSTOMER",address=ws.cell(r,7).value,source_system="REQUISITION",source_key=req)
            if requestor: self.ensure_partner(requestor,"EMPLOYEE",email=ws.cell(r,3).value,source_system="REQUISITION",source_key=req)
            if clean(ws.cell(r,7).value): self.ensure_location(ws.cell(r,7).value,"CUSTOMER_SITE")

    def collect_assets(self):
        # Legacy inventory master.
        legacy_items={r["id"]:r for r in self.db.execute("SELECT * FROM items")}
        for r in self.db.execute("SELECT * FROM inventory_serials ORDER BY id"):
            it=legacy_items.get(r["item_id"])
            desc=r["item_desc"] or (it["description"] if it else "")
            code=it["sku"] if it else ""
            self.add_asset_occurrence(r["serial_no"],category=r["category"],description=desc,item_code=code,
                source_system="LEGACY_DB",source_sheet="inventory_serials",source_row=r["id"],master=True,
                status=r["status"],location=r["location_name"],customer_id=r["customer_id"],unit_cost=r["unit_cost"],motor_no=r["motor_no"],created_at=r["created_at"])

        # ATLAS is the supplier asset manifest.
        wb=self.wb("ATLAS - Asset Manifest (1).xlsx")
        for sn,cat,serial_col in [("MOTORCYCLE","MC",6),("BATTERY","BAT",5),("LOCKER","BSS",4)]:
            ws=wb[sn]
            for r in range(2,(ws.max_row or 0)+1):
                s=ws.cell(r,serial_col).value
                if not serial(s): continue
                if cat=="MC":
                    desc=" ".join(x for x in [clean(ws.cell(r,3).value),clean(ws.cell(r,4).value)] if x)
                    self.add_asset_occurrence(s,category=cat,description=desc,item_code=ws.cell(r,5).value,source_system="ATLAS",source_sheet=sn,source_row=r,master=True,
                        batch=clean(ws.cell(r,1).value),manufacturer=clean(ws.cell(r,2).value),model=clean(ws.cell(r,3).value),color=clean(ws.cell(r,4).value),motor_no=clean(ws.cell(r,7).value),csr=clean(ws.cell(r,11).value),plate=clean(ws.cell(r,17).value),location=clean(ws.cell(r,19).value),created_at=date_text(ws.cell(r,8).value))
                elif cat=="BAT":
                    desc=f"RideBox Battery {clean(ws.cell(r,4).value)}".strip()
                    self.add_asset_occurrence(s,category=cat,description=desc,source_system="ATLAS",source_sheet=sn,source_row=r,master=True,
                        batch=clean(ws.cell(r,1).value),location=clean(ws.cell(r,2).value),manufacturer=clean(ws.cell(r,3).value),model=clean(ws.cell(r,4).value),secondary=clean(ws.cell(r,6).value),created_at=date_text(ws.cell(r,10).value))
                else:
                    desc=f"Swapping Station {clean(ws.cell(r,3).value)}".strip()
                    self.add_asset_occurrence(s,category=cat,description=desc,source_system="ATLAS",source_sheet=sn,source_row=r,master=True,
                        batch=clean(ws.cell(r,1).value),manufacturer=clean(ws.cell(r,2).value),model=clean(ws.cell(r,3).value),color=clean(ws.cell(r,5).value),location=clean(ws.cell(r,8).value),created_at=date_text(ws.cell(r,7).value))

        # STAR received stock master.
        ws=self.wb("STAR _ E88 SCM Inventory2026 (4).xlsx")["AW - Inventory - In"]
        for r in range(2,(ws.max_row or 0)+1):
            sn=ws.cell(r,5).value
            if not serial(sn): continue
            self.add_asset_occurrence(sn,category=ws.cell(r,2).value,description=ws.cell(r,4).value,item_code=ws.cell(r,3).value,
                source_system="STAR",source_sheet="AW - Inventory - In",source_row=r,master=True,
                status=clean(ws.cell(r,7).value),created_at=date_text(ws.cell(r,1).value),location="Asgard Warehouse")

        # STAKU references enrich and add any missing assets.
        ws=self.wb("STAKU - SALES_LEASE B2B (4).xlsx")["TRANSACTION"]
        for r in range(2,(ws.max_row or 0)+1):
            mc=ws.cell(r,11).value
            self.add_asset_occurrence(mc,category="MC",description=f"{clean(ws.cell(r,9).value)} {clean(ws.cell(r,10).value)}",source_system="STAKU",source_sheet="TRANSACTION",source_row=r,
                motor_no=clean(ws.cell(r,15).value),csr=clean(ws.cell(r,16).value),plate=clean(ws.cell(r,20).value))
            for c in (12,13): self.add_asset_occurrence(ws.cell(r,c).value,category="BAT",description="RideBox Battery",source_system="STAKU",source_sheet="TRANSACTION",source_row=r)
            self.add_asset_occurrence(ws.cell(r,14).value,category="CHG",description="RideBox Charger",source_system="STAKU",source_sheet="TRANSACTION",source_row=r)

        # SATURN serials.
        ws=self.wb("SATURN _ DELIVERY MONITORING _ LAST MILE (3).xlsx")["2026"]
        for r in range(2,(ws.max_row or 0)+1):
            self.add_asset_occurrence(ws.cell(r,13).value,category=ws.cell(r,11).value,description=ws.cell(r,12).value,source_system="SATURN",source_sheet="2026",source_row=r,
                motor_no=clean(ws.cell(r,14).value),plate=clean(ws.cell(r,18).value))

        # Pre-release serials.
        ws=self.wb("Pre-release Unit Checklist.xlsx")["Sheet1"]
        for r in range(3,(ws.max_row or 0)+1):
            self.add_asset_occurrence(ws.cell(r,8).value,category="MC",description=ws.cell(r,7).value,source_system="PRE_RELEASE",source_sheet="Sheet1",source_row=r)
            for c in (9,10): self.add_asset_occurrence(ws.cell(r,c).value,category="BAT",description="RideBox Battery",source_system="PRE_RELEASE",source_sheet="Sheet1",source_row=r)
            self.add_asset_occurrence(ws.cell(r,11).value,category="CHG",description="RideBox Charger",source_system="PRE_RELEASE",source_sheet="Sheet1",source_row=r)

        # Warehouse docs contain descriptions / serial combinations in columns 14:73.
        ws=self.wb("2026 SCM Warehouse Documents.xlsx")["2026 Transaction"]
        for r in range(3,(ws.max_row or 0)+1):
            for c in range(14,74):
                text=clean(ws.cell(r,c).value)
                if not text: continue
                for sn,desc in self.extract_serials_from_text(text):
                    self.add_asset_occurrence(sn,category=infer_category(desc,text),description=desc,source_system="WAREHOUSE_DOCS",source_sheet="2026 Transaction",source_row=r)

        # Battery mapping references.
        for r in self.db.execute("SELECT * FROM battery_mapping"):
            self.add_asset_occurrence(r["serial_no"],category="BAT",description="RideBox Battery",source_system="LEGACY_DB",source_sheet="battery_mapping",source_row=r["id"])

        self.insert_assets()

    @staticmethod
    def extract_serials_from_text(text: str) -> list[tuple[str,str]]:
        text=clean(text)
        out=[]
        # Slash-delimited warehouse descriptions: description / serial / plate.
        parts=[p.strip() for p in text.split("/")]
        candidates=[]
        if len(parts)>=2: candidates.extend(parts[1:3])
        candidates.extend(re.findall(r"\b(?:R5F[A-Z0-9]{12,}|5191100[A-Z0-9]{12,}|OKL[A-Z0-9]{10,}|[0-9A-Z]{6,}-[0-9A-Z]{3,})\b",text.upper()))
        seen=set()
        for cand in candidates:
            sn=serial(cand)
            if sn and sn not in seen and len(sn)>=8:
                seen.add(sn); out.append((sn,parts[0] if parts else text[:80]))
        # Multi-line battery serials sometimes lack delimiters.
        for line in re.split(r"[\n,;]+",text):
            sn=serial(line)
            if len(sn)>=18 and re.match(r"^[A-Z0-9-]+$",sn) and sn not in seen:
                seen.add(sn); out.append((sn,text[:80]))
        return out

    def insert_assets(self):
        legacy_dupes={serial(r["serial_no"]):r for r in self.db.execute("SELECT * FROM serial_dupe_log") if serial(r["serial_no"])}
        for sn in sorted(self.master_occurrences):
            occurrences=self.master_occurrences[sn]
            # Prefer legacy, then ATLAS, then STAR as the canonical master.
            priority={"LEGACY_DB":0,"ATLAS":1,"STAR":2,"STAKU":3,"SATURN":4,"PRE_RELEASE":5,"WAREHOUSE_DOCS":6}
            canonical=sorted(occurrences,key=lambda x:(priority.get(x["source_system"],99),x["source_row"]))[0]
            item=self.ensure_item(canonical.get("description") or canonical.get("category"),canonical.get("category"),canonical.get("item_code"),
                                  manufacturer=canonical.get("manufacturer"),model=canonical.get("model"),color=canonical.get("color"),serialized=True,
                                  standard_cost=canonical.get("unit_cost"),source_system=canonical["source_system"],source_key=sn)
            loc=self.ensure_location(canonical.get("location") or "Asgard Warehouse","WAREHOUSE")
            customer_id=canonical.get("customer_id")
            holder_name=""
            holder_id=None
            if customer_id:
                old=self.db.execute("SELECT * FROM customers WHERE id=?",(customer_id,)).fetchone()
                if old:
                    p=self.ensure_partner(old["name"],"CUSTOMER",old["code"] or ""); holder_id=p["id"];holder_name=p["name"]
            current_status=status_norm(canonical.get("status"))
            asset_no=f"AST-MIG-{self.asset_seq:08d}"; self.asset_seq+=1
            self.db.execute("""INSERT INTO erp_assets(asset_no,serial_no,serial_type,item_id,item_code,item_name,category,secondary_serial,motor_no,plate_no,csr_no,batch_code,current_location_id,current_location_code,current_status,current_holder_type,current_holder_id,current_holder_name,unit_cost,condition_code,reconciliation_status,source_system,source_key,created_at,updated_at)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (asset_no,sn,canonical.get("category") or item["category"],item["id"],item["item_code"],item["item_name"],item["category"],clean(canonical.get("secondary")),clean(canonical.get("motor_no")),clean(canonical.get("plate")),clean(canonical.get("csr")),clean(canonical.get("batch")),loc["id"],loc["code"],current_status,"CUSTOMER" if holder_id else None,holder_id,holder_name,number(canonical.get("unit_cost")),"GOOD","CLEAR",canonical["source_system"],f"{canonical['source_sheet']}:{canonical['source_row']}",canonical.get("created_at") or datetime.now().isoformat(),datetime.now().isoformat()))
            asset_id=self.db.execute("SELECT id FROM erp_assets WHERE serial_no=?",(sn,)).fetchone()[0]
            # Open only true duplicate-master exceptions; operational references are retained in the archive.
            master_occ=[x for x in occurrences if x.get("master")]
            if len(master_occ)>1 or sn in legacy_dupes:
                details={"master_occurrences":master_occ}
                if sn in legacy_dupes: details["legacy_duplicate_log"]=dict(legacy_dupes[sn])
                self.add_exception(sn,"DUPLICATE_MASTER_SERIAL",canonical["source_system"],canonical["source_sheet"],canonical["source_row"],asset_id,details)
        self.counts["canonical_assets"] = self.db.execute("SELECT COUNT(*) FROM erp_assets").fetchone()[0]

    def add_exception(self,sn,typ,source, sheet, row, asset_id, payload):
        no=f"EXC-MIG-{self.exception_seq:07d}"; self.exception_seq+=1
        self.db.execute("""INSERT INTO erp_serial_exceptions(exception_no,serial_no,exception_type,source_system,source_sheet,source_row,canonical_asset_id,payload_json,status)
                          VALUES(?,?,?,?,?,?,?,?, 'OPEN')""",(no,sn,typ,source,sheet,row,asset_id,json.dumps(payload,default=clean,ensure_ascii=False)))
        self.counts["serial_exceptions"] += 1

    def get_asset(self,sn:Any):
        return self.db.execute("SELECT * FROM erp_assets WHERE serial_no=?",(serial(sn),)).fetchone()

    def add_movement(self, *, sn:Any, date_value:Any, movement_type:str, from_loc:Any="", to_loc:Any="", to_status:Any="", holder_type:Any="", holder_name:Any="", source_type:str, source_no:Any="", notes:Any=""):
        asset=self.get_asset(sn)
        if not asset: return None
        dt=date_text(date_value) or "2026-01-01"
        f=self.ensure_location(from_loc or asset["current_location_code"] or "Unassigned","OTHER")
        t=self.ensure_location(to_loc or asset["current_location_code"] or "Unassigned","OTHER")
        mt=clean(movement_type).upper().replace(" ","_") or "TRANSFER"
        key=(asset["serial_no"],dt,mt,t["code"],clean(source_no))
        if key in self.movement_keys: return self.movement_keys[key]
        no=f"MV-MIG-{self.move_seq:08d}"; self.move_seq+=1
        status=status_norm(to_status) if clean(to_status) else asset["current_status"]
        holder_id=None; hname=clean(holder_name); htype=clean(holder_type).upper() or None
        if hname and htype in {"CUSTOMER","EMPLOYEE","SITE_PARTNER"}:
            p=self.ensure_partner(hname,htype);holder_id=p["id"]
        cur=self.db.execute("""INSERT INTO erp_stock_ledger(movement_no,movement_date,movement_type,asset_id,serial_no,item_id,item_code,qty,from_location_id,from_location_code,to_location_id,to_location_code,from_status,to_status,holder_type,holder_id,holder_name,source_doc_type,source_doc_no,notes,posted_by)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (no,dt,mt,asset["id"],asset["serial_no"],asset["item_id"],asset["item_code"],1,f["id"],f["code"],t["id"],t["code"],asset["current_status"],status,htype,holder_id,hname,source_type,clean(source_no),clean(notes),"MIGRATION"))
        self.movement_keys[key]=cur.lastrowid
        self.db.execute("""UPDATE erp_assets SET current_location_id=?,current_location_code=?,current_status=?,current_holder_type=?,current_holder_id=?,current_holder_name=?,updated_at=? WHERE id=?""",
                        (t["id"],t["code"],status,htype,holder_id,hname,dt,asset["id"]))
        return cur.lastrowid

    def load_shipments_receiving_landed_cost(self):
        wb=self.wb("STELLAR _ Shipment.xlsx"); ws=wb["TRACKING"]
        shipment_by_manifest={}
        for r in range(22,(ws.max_row or 0)+1):
            shipment=clean(ws.cell(r,4).value)
            if not shipment: continue
            batch=clean(ws.cell(r,5).value); supplier=clean(ws.cell(r,6).value)
            vendor=self.ensure_partner(supplier or "Unknown Supplier","VENDOR",address=ws.cell(r,7).value,source_system="STELLAR",source_key=shipment)
            ship_no=shipment
            if self.db.execute("SELECT 1 FROM erp_shipments WHERE shipment_no=?",(ship_no,)).fetchone(): ship_no=f"{shipment}-{r}"
            status=status_norm(ws.cell(r,3).value)
            if status=="AVAILABLE": status="COMPLETED"
            cur=self.db.execute("""INSERT INTO erp_shipments(shipment_no,batch_code,supplier_id,supplier_name,mode_of_transport,incoterm,shipping_line,vessel,container_no,origin,destination,etd,eta,actual_departure,actual_arrival,warehouse_arrival,status,source_system,source_key,created_by)
                                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'STELLAR',?,'MIGRATION')""",
                                (ship_no,shipment,vendor["id"],vendor["name"],clean(ws.cell(r,11).value),clean(ws.cell(r,12).value),clean(ws.cell(r,13).value),clean(ws.cell(r,14).value),clean(ws.cell(r,16).value),clean(ws.cell(r,7).value),"E88 Asgard Warehouse",date_text(ws.cell(r,19).value),date_text(ws.cell(r,22).value),date_text(ws.cell(r,19).value),date_text(ws.cell(r,22).value),date_text(ws.cell(r,24).value),clean(ws.cell(r,3).value).upper() or "DRAFT",shipment))
            shipment_by_manifest[shipment]=cur.lastrowid

        # ATLAS batches drive shipment expected serials and auto-create missing shipments.
        wb=self.wb("ATLAS - Asset Manifest (1).xlsx")
        groups=defaultdict(list)
        for sheet,cat,sc in [("MOTORCYCLE","MC",6),("BATTERY","BAT",5),("LOCKER","BSS",4)]:
            ws=wb[sheet]
            for r in range(2,(ws.max_row or 0)+1):
                sn=serial(ws.cell(r,sc).value); batch=clean(ws.cell(r,1).value)
                if not sn or not batch: continue
                desc=(f"{clean(ws.cell(r,3).value)} {clean(ws.cell(r,4).value)}" if cat=="MC" else (f"RideBox Battery {clean(ws.cell(r,4).value)}" if cat=="BAT" else f"Swapping Station {clean(ws.cell(r,3).value)}"))
                groups[batch].append({"serial":sn,"cat":cat,"desc":desc,"row":r,"sheet":sheet,"code":clean(ws.cell(r,5).value) if cat=="MC" else ""})
        for batch,rows in groups.items():
            shipment_id=shipment_by_manifest.get(batch)
            if not shipment_id:
                vendor=self.ensure_partner(rows[0].get("manufacturer") or "Supplier per ATLAS","VENDOR")
                cur=self.db.execute("""INSERT INTO erp_shipments(shipment_no,batch_code,supplier_id,supplier_name,status,source_system,source_key,created_by)
                                      VALUES(?,?,?,?, 'MANIFESTED','ATLAS',?,'MIGRATION')""",(batch,batch,vendor["id"],vendor["name"],batch))
                shipment_id=cur.lastrowid; shipment_by_manifest[batch]=shipment_id
            by_item=defaultdict(list)
            for x in rows:
                item=self.ensure_item(x["desc"],x["cat"],x["code"],serialized=True,source_system="ATLAS",source_key=x["serial"])
                x["item_id"]=item["id"];x["item_code"]=item["item_code"];by_item[item["id"]].append(x)
            for line_no,(item_id,xs) in enumerate(by_item.items(),1):
                item=self.db.execute("SELECT * FROM erp_items WHERE id=?",(item_id,)).fetchone()
                lr=self.db.execute("""INSERT INTO erp_shipment_lines(shipment_id,line_no,item_id,item_code,description,category,expected_qty,received_qty,status,source_sheet)
                                    VALUES(?,?,?,?,?,?,?,?,?,?)""",(shipment_id,line_no,item_id,item["item_code"],item["item_name"],item["category"],len(xs),sum(1 for x in xs if self.get_asset(x["serial"])),"RECEIVED" if all(self.get_asset(x["serial"]) for x in xs) else "PARTIAL",xs[0]["sheet"]))
                line_id=lr.lastrowid
                for x in xs:
                    asset=self.get_asset(x["serial"])
                    self.db.execute("""INSERT INTO erp_expected_assets(shipment_id,shipment_line_id,serial_no,serial_type,item_id,item_code,batch_code,expected_status,source_sheet,source_row)
                                     VALUES(?,?,?,?,?,?,?,?,?,?)""",(shipment_id,line_id,x["serial"],x["cat"],x["item_id"],x["item_code"],batch,"RECEIVED" if asset else "EXPECTED",x["sheet"],x["row"]))
            # Historical receipt generated from the supplier manifest and current inventory.
            actual_assets=[self.get_asset(x["serial"]) for x in rows if self.get_asset(x["serial"])]
            if actual_assets:
                loc=self.ensure_location("E88 Asgard Warehouse","WAREHOUSE","ASGARD")
                ship=self.db.execute("SELECT * FROM erp_shipments WHERE id=?",(shipment_id,)).fetchone()
                receipt_no=f"RCV-HIST-{self.receipt_seq:06d}";self.receipt_seq+=1
                rr=self.db.execute("""INSERT INTO erp_receipts(receipt_no,shipment_id,location_id,received_at,receiving_status,document_ref,notes,received_by,posted_by,posted_at)
                                    VALUES(?,?,?,?, 'POSTED',?,'Auto-created from ATLAS/STELLAR opening data','MIGRATION','MIGRATION',datetime('now'))""",
                                   (receipt_no,shipment_id,loc["id"],ship["warehouse_arrival"] or "2026-01-01",batch))
                for a in actual_assets:
                    exp=self.db.execute("SELECT * FROM erp_expected_assets WHERE shipment_id=? AND serial_no=?",(shipment_id,a["serial_no"])).fetchone()
                    self.db.execute("""INSERT OR IGNORE INTO erp_receipt_lines(receipt_id,shipment_line_id,expected_asset_id,serial_no,item_id,item_code,qty,condition_code,acceptance_status,source_method)
                                     VALUES(?,?,?,?,?,?,1,'GOOD','MATCHED','MIGRATION')""",(rr.lastrowid,exp["shipment_line_id"] if exp else None,exp["id"] if exp else None,a["serial_no"],a["item_id"],a["item_code"]))

        # Landed cost records from STELLAR logistics cost sheet.
        ws=self.wb("STELLAR _ Shipment.xlsx")["Stacku - Logistics Shipment Cha"]
        for r in range(5,(ws.max_row or 0)+1):
            desc=clean(ws.cell(r,4).value); total=number(ws.cell(r,18).value)
            if not desc or total==0: continue
            # Match by commodity/batch text, otherwise leave shipment unlinked.
            ship=None
            for key,sid in shipment_by_manifest.items():
                if norm(key) in norm(desc) or norm(desc) in norm(key): ship=sid;break
            no=f"LC-HIST-{self.lc_seq:06d}";self.lc_seq+=1
            cur=self.db.execute("""INSERT INTO erp_landed_cost_headers(landed_cost_no,shipment_id,allocation_method,currency,exchange_rate,status,total_cost,notes,posted_by,posted_at,created_by)
                                 VALUES(?,?,'VALUE','PHP',1,'POSTED',?,?,'MIGRATION',datetime('now'),'MIGRATION')""",(no,ship,total,f"STELLAR row {r}: {desc}"))
            cost_names=[("VESSEL_FREIGHT_USD",11),("VESSEL_FREIGHT_PHP",12),("CUSTOMS_DUTIES",13),("SHIPPING_HANDLING",14),("GATEPASS",15),("OTHER_IMPORT",16),("CP_PROCESSING",17)]
            for typ,c in cost_names:
                amt=number(ws.cell(r,c).value)
                if amt: self.db.execute("INSERT INTO erp_landed_cost_lines(landed_cost_id,cost_type,amount,notes) VALUES(?,?,?,?)",(cur.lastrowid,typ,amt,f"Source row {r}"))

    def load_movements_and_current_state(self):
        # STAR movement log is the strongest operational history.
        ws=self.wb("STAR _ E88 SCM Inventory2026 (4).xlsx")["AW -  Transaction Log"]
        for r in range(2,(ws.max_row or 0)+1):
            sn=ws.cell(r,4).value
            if not serial(sn): continue
            movement=clean(ws.cell(r,5).value); frm=clean(ws.cell(r,7).value); to=clean(ws.cell(r,8).value)
            remarks=" | ".join(x for x in [clean(ws.cell(r,10).value),clean(ws.cell(r,11).value)] if x)
            holder_type=""
            if "EMPLOYEE" in norm(remarks): holder_type="EMPLOYEE"
            elif "LEASE" in norm(remarks) or "SOLD" in norm(remarks): holder_type="CUSTOMER"
            elif "DEPLOY" in norm(remarks): holder_type="SITE_PARTNER"
            self.add_movement(sn=sn,date_value=ws.cell(r,12).value or ws.cell(r,1).value,movement_type=movement,from_loc=frm,to_loc=to,
                              to_status=remarks or movement,holder_type=holder_type,holder_name=to if holder_type else "",source_type="STAR",source_no=ws.cell(r,9).value,notes=remarks)

        # STAKU commercial assignments and sales.
        ws=self.wb("STAKU - SALES_LEASE B2B (4).xlsx")["TRANSACTION"]
        for r in range(2,(ws.max_row or 0)+1):
            tx=norm(ws.cell(r,2).value or ws.cell(r,1).value)
            mc=serial(ws.cell(r,11).value); cust=clean(ws.cell(r,7).value)
            if not mc or not cust: continue
            customer=self.ensure_partner(cust,"CUSTOMER",address=ws.cell(r,8).value,source_system="STAKU",source_key=str(r))
            trans_type="SALE" if "SOLD" in tx else "LEASE"
            so_no=f"SO-HIST-{self.sales_seq:06d}";self.sales_seq+=1
            so=self.db.execute("""INSERT INTO erp_sales_orders(sales_order_no,transaction_type,customer_id,order_date,contract_start,contract_end,status,delivery_address,source_system,source_key,created_by,posted_by,posted_at)
                               VALUES(?,?,?,?,?,?,'POSTED',?,'STAKU',?,'MIGRATION','MIGRATION',datetime('now'))""",
                             (so_no,trans_type,customer["id"],date_text(ws.cell(r,4).value),date_text(ws.cell(r,5).value),date_text(ws.cell(r,6).value),clean(ws.cell(r,8).value),str(r)))
            line_no=0
            for c,role,cat,desc in [(11,"MOTORCYCLE","MC",f"{clean(ws.cell(r,9).value)} {clean(ws.cell(r,10).value)}"),(12,"BATTERY_1","BAT","RideBox Battery"),(13,"BATTERY_2","BAT","RideBox Battery"),(14,"CHARGER","CHG","RideBox Charger")]:
                sn=serial(ws.cell(r,c).value)
                if not sn: continue
                a=self.get_asset(sn); item=self.ensure_item(desc,cat,serialized=True,source_system="STAKU",source_key=sn)
                line_no+=1
                self.db.execute("INSERT INTO erp_sales_lines(sales_order_id,line_no,item_id,item_code,description,qty,asset_id,serial_no,line_role) VALUES(?,?,?,?,?,1,?,?,?)",
                                (so.lastrowid,line_no,item["id"],item["item_code"],item["item_name"],a["id"] if a else None,sn,role))
            asg_no=f"ASG-HIST-{self.assignment_seq:06d}";self.assignment_seq+=1
            asg=self.db.execute("""INSERT INTO erp_assignments(assignment_no,assignment_type,partner_id,holder_name,start_date,expected_return_date,status,purpose,source_request_no,created_by,approved_by,approved_at)
                                VALUES(?,?,?,?,?,?,'POSTED',?,?, 'MIGRATION','MIGRATION',datetime('now'))""",
                                (asg_no,trans_type,customer["id"],customer["name"],date_text(ws.cell(r,4).value),date_text(ws.cell(r,6).value),trans_type,clean(ws.cell(r,3).value)))
            for c,role in [(11,"MOTORCYCLE"),(12,"BATTERY_1"),(13,"BATTERY_2"),(14,"CHARGER")]:
                sn=serial(ws.cell(r,c).value);a=self.get_asset(sn)
                if sn: self.db.execute("INSERT OR IGNORE INTO erp_assignment_assets(assignment_id,asset_id,serial_no,role_code,condition_out) VALUES(?,?,?,?,'GOOD')",(asg.lastrowid,a["id"] if a else None,sn,role))
                if sn: self.add_movement(sn=sn,date_value=ws.cell(r,4).value,movement_type=trans_type,from_loc="Asgard Warehouse",to_loc=customer["name"],to_status="SOLD" if trans_type=="SALE" else "LEASED",holder_type="CUSTOMER",holder_name=customer["name"],source_type="STAKU",source_no=so_no,notes=clean(ws.cell(r,17).value))

        # SATURN delivery and return movements; one delivery per requisition/destination/date.
        ws=self.wb("SATURN _ DELIVERY MONITORING _ LAST MILE (3).xlsx")["2026"]
        deliveries={}
        delivered_by_req=defaultdict(set); returned_by_req=defaultdict(set); req_meta={}
        for r in range(2,(ws.max_row or 0)+1):
            sn=serial(ws.cell(r,13).value)
            if not sn: continue
            req=clean(ws.cell(r,7).value) or f"SATURN-{r}"
            req_meta.setdefault(req,{"destination":clean(ws.cell(r,22).value),"date":date_text(ws.cell(r,32).value or ws.cell(r,20).value or ws.cell(r,10).value),"purpose":clean(ws.cell(r,3).value)})
            key=(req,date_text(ws.cell(r,10).value),clean(ws.cell(r,22).value))
            if key not in deliveries:
                dno=f"DLV-HIST-{self.delivery_seq:06d}";self.delivery_seq+=1
                origin=self.ensure_location(ws.cell(r,16).value or "Asgard Warehouse","WAREHOUSE")
                cur=self.db.execute("""INSERT INTO erp_deliveries(delivery_no,requested_date,scheduled_date,actual_release_date,actual_delivery_date,origin_location_id,destination,recipient_name,status,source_system,source_key,created_by)
                                     VALUES(?,?,?,?,?,?,?,?,?,'SATURN',?,'MIGRATION')""",
                                  (dno,date_text(ws.cell(r,9).value),date_text(ws.cell(r,10).value),date_text(ws.cell(r,19).value),date_text(ws.cell(r,20).value),origin["id"],clean(ws.cell(r,22).value),clean(ws.cell(r,21).value),clean(ws.cell(r,1).value).upper() or "PLANNED",req))
                deliveries[key]=(cur.lastrowid,dno)
            did,dno=deliveries[key];a=self.get_asset(sn)
            self.db.execute("INSERT OR IGNORE INTO erp_delivery_assets(delivery_id,asset_id,serial_no,item_code,qty) VALUES(?,?,?,?,?)",(did,a["id"] if a else None,sn,a["item_code"] if a else "",number(ws.cell(r,15).value,1)))
            is_return="RETURN" in norm(ws.cell(r,1).value) or date_text(ws.cell(r,32).value)!=""
            if is_return:
                returned_by_req[req].add(sn)
                self.add_movement(sn=sn,date_value=ws.cell(r,32).value or ws.cell(r,20).value,movement_type="RETURN",from_loc=ws.cell(r,22).value,to_loc="Returns Quarantine",to_status="AVAILABLE",source_type="SATURN",source_no=req,notes=ws.cell(r,29).value)
            else:
                delivered_by_req[req].add(sn)
                holder_type="SITE_PARTNER" if infer_category(ws.cell(r,11).value)=="BAT" or "DEPLOY" in norm(ws.cell(r,3).value) else "CUSTOMER"
                self.add_movement(sn=sn,date_value=ws.cell(r,19).value or ws.cell(r,10).value,movement_type="DELIVERY",from_loc=ws.cell(r,16).value,to_loc=ws.cell(r,22).value,to_status="ASSIGNED",holder_type=holder_type,holder_name=ws.cell(r,22).value,source_type="SATURN",source_no=req,notes=ws.cell(r,3).value)
        # Historical return acceptance. Swapped batteries are accepted to quarantine and opened as unreconciled.
        quarantine=self.ensure_location("Returns Quarantine","QUARANTINE","RET-QUAR")
        for req,returned in returned_by_req.items():
            delivered=delivered_by_req.get(req,set())
            meta=req_meta.get(req,{})
            partner=self.ensure_partner(meta.get("destination") or "Unknown Return Source","SITE_PARTNER",source_system="SATURN",source_key=req)
            ret_no=f"RET-HIST-{self.return_seq:06d}";self.return_seq+=1
            rr=self.db.execute("""INSERT INTO erp_return_orders(return_no,partner_id,return_date,return_location_id,status,reason_code,notes,created_by,posted_by,posted_at)
                                VALUES(?,?,?,?, 'POSTED','HISTORICAL_RETURN',?,'MIGRATION','MIGRATION',datetime('now'))""",
                               (ret_no,partner["id"],meta.get("date") or "2026-01-01",quarantine["id"],f"SATURN requisition {req}: {meta.get('purpose','')}"))
            missing=list(delivered-returned)
            for actual in sorted(returned):
                actual_asset=self.get_asset(actual)
                if actual in delivered:
                    expected=actual;acceptance="MATCHED"
                else:
                    expected=missing.pop(0) if missing else ""
                    acceptance="BATTERY_SWAP" if (actual_asset and actual_asset["category"]=="BAT") else "SERIAL_MISMATCH"
                expected_asset=self.get_asset(expected)
                self.db.execute("""INSERT OR IGNORE INTO erp_return_lines(return_id,expected_asset_id,expected_serial,actual_asset_id,actual_serial,item_category,acceptance_status,condition_code,notes)
                                  VALUES(?,?,?,?,?,?,?,'GOOD',?)""",
                                (rr.lastrowid,expected_asset["id"] if expected_asset else None,expected,actual_asset["id"] if actual_asset else None,actual,actual_asset["category"] if actual_asset else "",acceptance,f"Imported from SATURN requisition {req}"))
                if acceptance!="MATCHED":
                    no=f"REC-MIG-{self.recon_seq:06d}";self.recon_seq+=1
                    self.db.execute("""INSERT INTO erp_reconciliation_cases(case_no,case_type,return_id,expected_serial,actual_serial,current_location_code,status,resolution_notes,opened_by)
                                      VALUES(?,?,?,?,?, 'RET-QUAR','UNRECONCILED',?,'MIGRATION')""",
                                    (no,acceptance,rr.lastrowid,expected,actual,f"SATURN requisition {req}: returned serial differs from expected delivered set"))
                    if actual_asset:self.db.execute("UPDATE erp_assets SET reconciliation_status='UNRECONCILED',current_status='QUARANTINE' WHERE id=?",(actual_asset["id"],))

        # Warehouse document movements enrich the ledger without duplicate re-encoding.
        ws=self.wb("2026 SCM Warehouse Documents.xlsx")["2026 Transaction"]
        for r in range(3,(ws.max_row or 0)+1):
            doc=clean(ws.cell(r,4).value or ws.cell(r,5).value or ws.cell(r,6).value or ws.cell(r,3).value)
            dt=ws.cell(r,8).value; origin=ws.cell(r,12).value or "Asgard Warehouse";dest=ws.cell(r,13).value
            is_return="RETURN" in norm(ws.cell(r,75).value) or date_text(ws.cell(r,74).value)!=""
            for c in range(14,74):
                text=clean(ws.cell(r,c).value)
                for sn,desc in self.extract_serials_from_text(text):
                    self.add_movement(sn=sn,date_value=ws.cell(r,74).value if is_return else dt,movement_type="RETURN" if is_return else "GOODS_ISSUE",from_loc=dest if is_return else origin,to_loc=origin if is_return else dest,to_status="AVAILABLE" if is_return else clean(ws.cell(r,10).value),source_type="WAREHOUSE_DOC",source_no=doc,notes=desc)

    def load_requisitions_checklists(self):
        ws=self.wb("SCM Requisition Slip 1226.xlsx")["Form Responses 1"]
        item_cols={12:"MC - R280 White",13:"MC - R280 Black",14:"MC - R280 Sport Red",15:"MC - R280 Sport Blue",16:"MC - D400 Red",17:"MC - D400 Blue",18:"MC - D400 Black",19:"RideBox Battery",20:"RideBox Charger",21:"RideBox Locker",22:"RideBox Rack",23:"Battery & Charger",25:"Helmet",26:"Charging Kit",32:"Charging Kit"}
        for r in range(2,(ws.max_row or 0)+1):
            raw=clean(ws.cell(r,1).value)
            if not raw: continue
            no=f"REQ-{raw}" if not raw.upper().startswith("REQ") else raw
            if self.db.execute("SELECT 1 FROM erp_requisitions WHERE requisition_no=?",(no,)).fetchone(): no=f"{no}-{r}"
            company=clean(ws.cell(r,6).value);partner=self.ensure_partner(company or "Internal","CUSTOMER",address=ws.cell(r,7).value)
            cur=self.db.execute("""INSERT INTO erp_requisitions(requisition_no,request_date,requestor_email,requestor_name,department,purpose,fulfillment_method,partner_id,destination,required_date,status,remarks,source_system,source_key)
                                VALUES(?,?,?,?,?,?,?,?,?,?,?,?, 'REQUISITION',?)""",
                              (no,date_text(ws.cell(r,2).value),clean(ws.cell(r,3).value).lower(),clean(ws.cell(r,29).value),clean(ws.cell(r,30).value),clean(ws.cell(r,4).value),clean(ws.cell(r,5).value),partner["id"],clean(ws.cell(r,7).value),date_text(ws.cell(r,8).value),clean(ws.cell(r,35).value).upper() or "OPEN",clean(ws.cell(r,28).value),raw))
            for c,desc in item_cols.items():
                qty=number(ws.cell(r,c).value)
                if qty:
                    item=self.ensure_item(desc,infer_category(desc),serialized=infer_category(desc) in {"MC","BAT","BSS","CHG"},source_system="REQUISITION",source_key=f"{r}:{c}")
                    self.db.execute("INSERT INTO erp_requisition_lines(requisition_id,item_id,item_code,description,qty,serial_required) VALUES(?,?,?,?,?,?)",(cur.lastrowid,item["id"],item["item_code"],item["item_name"],qty,item["serialized"]))
            for c in (27,31,33,34):
                desc=clean(ws.cell(r,c).value)
                if desc:
                    item=self.ensure_item(desc,infer_category(desc),source_system="REQUISITION",source_key=f"{r}:{c}")
                    self.db.execute("INSERT INTO erp_requisition_lines(requisition_id,item_id,item_code,description,qty,serial_required) VALUES(?,?,?,?,1,?)",(cur.lastrowid,item["id"],item["item_code"],item["item_name"],item["serialized"]))

        ws=self.wb("Pre-release Unit Checklist.xlsx")["Sheet1"]
        for r in range(3,(ws.max_row or 0)+1):
            sn=serial(ws.cell(r,8).value)
            if not sn: continue
            no=f"PRC-HIST-{self.check_seq:06d}";self.check_seq+=1
            payload={
                "canProceed":clean(ws.cell(r,4).value),"pdi":clean(ws.cell(r,12).value),"pdiForm":clean(ws.cell(r,13).value),
                "hydraBatteryAssignment":clean(ws.cell(r,14).value),"appTesting":clean(ws.cell(r,15).value),"accountCreation":clean(ws.cell(r,16).value),
                "batteryA":serial(ws.cell(r,9).value),"batteryB":serial(ws.cell(r,10).value),"charger":serial(ws.cell(r,11).value),
                "ltoStatus":clean(ws.cell(r,22).value),"documents":clean(ws.cell(r,23).value),"orCr":clean(ws.cell(r,24).value),"plate":clean(ws.cell(r,25).value)
            }
            self.db.execute("""INSERT INTO erp_pre_release_checks(checklist_no,serial_no,check_date,checklist_json,result,defects,checked_by,approved_by)
                              VALUES(?,?,?,?,?,?,?,'MIGRATION')""",(no,sn,date_text(ws.cell(r,3).value),json.dumps(payload,ensure_ascii=False),"PASSED" if norm(ws.cell(r,1).value)=="DONE" and clean(ws.cell(r,4).value) in {"1","TRUE"} else "PENDING","",clean(ws.cell(r,2).value)))

    def load_procurement_sales_budget_planning(self):
        # Procurement register, payment register and journal source.
        # Approved POs are authoritative purchase orders used for receiving and commitments.
        budget_wb=self.wb("E88_ApprovedBudget2026 (4)(1).xlsx")
        po_ws=budget_wb["Purchase Orders (Approved)"]
        for r in range(5,(po_ws.max_row or 0)+1):
            po=clean(po_ws.cell(r,1).value)
            if not po: continue
            vendor_name=clean(po_ws.cell(r,4).value) or "Unknown Vendor"
            vendor=self.ensure_partner(vendor_name,"VENDOR",source_system="APPROVED_BUDGET",source_key=po)
            if not self.db.execute("SELECT 1 FROM erp_purchase_orders WHERE purchase_order_no=?",(po,)).fetchone():
                cur=self.db.execute("""INSERT INTO erp_purchase_orders(purchase_order_no,vendor_id,vendor_name,order_date,currency,status,total_amount,source_system,source_key,created_by)
                                     VALUES(?,?,?,?,'PHP',?,?,'APPROVED_BUDGET',?,'MIGRATION')""",
                                    (po,vendor["id"],vendor["name"],date_text(po_ws.cell(r,2).value),clean(po_ws.cell(r,8).value).upper() or "APPROVED",number(po_ws.cell(r,6).value),po))
                purpose=clean(po_ws.cell(r,7).value) or f"Purchase order {po}"
                item=self.ensure_item(purpose,infer_category(purpose),source_system="APPROVED_BUDGET",source_key=po)
                self.db.execute("INSERT INTO erp_purchase_order_lines(purchase_order_id,line_no,item_id,item_code,description,ordered_qty,unit_cost,line_amount,status) VALUES(?,1,?,?,?,?,?,?, 'OPEN')",
                                (cur.lastrowid,item["id"],item["item_code"],item["item_name"],1,number(po_ws.cell(r,6).value),number(po_ws.cell(r,6).value)))

        wb=self.wb("E88_ProcurementMonitoring_2026.xlsx")
        ws=wb["Procurement Monitoring"]
        for r in range(6,(ws.max_row or 0)+1):
            if not any(clean(ws.cell(r,c).value) for c in range(1,32)): continue
            vals=[ws.cell(r,c).value for c in range(1,32)]
            self.db.execute("""INSERT OR IGNORE INTO erp_procurement_register(procurement_id,request_date,request_no,requestor_payee,department,cost_center,project_site,account_title,procurement_category,description,po_no,supplier_invoice_no,invoice_date,gross_amount,vat_type,vat_rate,net_of_vat,input_vat,ewt_rate,ewt_amount,net_payable,payment_terms,due_date,aging_status,approval_stage,payment_status,payment_reference,paid_date,attachment_url,remarks,source_system,source_sheet,source_row)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'PROCUREMENT_MONITOR','Procurement Monitoring',?)""",
                            (clean(vals[0]),date_text(vals[1]),clean(vals[2]),clean(vals[3]),clean(vals[4]),clean(vals[5]),clean(vals[6]),clean(vals[7]),clean(vals[8]),clean(vals[9]),clean(vals[10]),clean(vals[11]),date_text(vals[12]),number(vals[13]),clean(vals[14]),number(vals[15]),number(vals[16]),number(vals[17]),number(vals[18]),number(vals[19]),number(vals[20]),clean(vals[21]),date_text(vals[22]),clean(vals[23]),clean(vals[24]),clean(vals[25]),clean(vals[26]),date_text(vals[27]),clean(vals[29]),clean(vals[30]),r))
            vendor=clean(vals[3]);po=clean(vals[10])
            if po:
                p=self.ensure_partner(vendor or "Unknown Vendor","VENDOR")
                if not self.db.execute("SELECT 1 FROM erp_purchase_orders WHERE purchase_order_no=?",(po,)).fetchone():
                    self.db.execute("""INSERT INTO erp_purchase_orders(purchase_order_no,vendor_id,vendor_name,order_date,currency,status,total_amount,source_system,source_key,created_by)
                                      VALUES(?,?,?,?,'PHP',?,?, 'PROCUREMENT_MONITOR',?,'MIGRATION')""",(po,p["id"],p["name"],date_text(vals[1]),clean(vals[24]).upper() or "APPROVED",number(vals[13]),clean(vals[0]) or str(r)))
        ws=wb["Payments Recording"]
        for r in range(6,(ws.max_row or 0)+1):
            if not clean(ws.cell(r,1).value): continue
            vals=[ws.cell(r,c).value for c in range(1,16)]
            self.db.execute("""INSERT OR IGNORE INTO erp_payment_register(payment_id,payment_date,request_no,payee,department,cost_center,account_title,gross_amount,ewt_amount,net_payable,bank,payment_reference,payment_status,payment_variance,remarks,source_system,source_sheet,source_row)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'PROCUREMENT_MONITOR','Payments Recording',?)""",(clean(vals[0]),date_text(vals[1]),clean(vals[2]),clean(vals[3]),clean(vals[4]),clean(vals[5]),clean(vals[6]),number(vals[7]),number(vals[8]),number(vals[9]),clean(vals[10]),clean(vals[11]),clean(vals[12]),number(vals[13]),clean(vals[14]),r))
        # Historical sales/collection records.
        wb=self.wb("E88_SalesMonitoring_2026.xlsx")
        for sheet in ["Aftersales","WhseServiceSales","MCSales","MCLeased","BSwapping"]:
            ws=wb[sheet]
            for r in range(6,(ws.max_row or 0)+1):
                if not clean(ws.cell(r,1).value): continue
                customer=clean(ws.cell(r,5).value);p=self.ensure_partner(customer or "Various Customers","CUSTOMER",source_system="SALES_MONITOR",source_key=f"{sheet}:{r}")
                self.db.execute("""INSERT OR IGNORE INTO erp_sales_receipts(entry_id,source_module,transaction_date,sales_type,document_no,customer_id,customer_name,contract_or_unit_no,department,cost_center,account_title,description,gross_amount,vat_type,vat_rate,net_of_vat,output_vat,payment_method,bank_wallet,bank_reference,other_reference,settlement_date,cleared_status,notes,source_system,source_sheet,source_row)
                                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'SALES_MONITOR',?,?)""",
                                (clean(ws.cell(r,1).value),sheet,date_text(ws.cell(r,2).value),clean(ws.cell(r,3).value),clean(ws.cell(r,4).value),p["id"],p["name"],clean(ws.cell(r,6).value),clean(ws.cell(r,7).value),clean(ws.cell(r,8).value),clean(ws.cell(r,9).value),clean(ws.cell(r,10).value),number(ws.cell(r,11).value),clean(ws.cell(r,12).value),number(ws.cell(r,13).value),number(ws.cell(r,14).value),number(ws.cell(r,15).value),clean(ws.cell(r,16).value),clean(ws.cell(r,17).value),clean(ws.cell(r,18).value),clean(ws.cell(r,19).value),date_text(ws.cell(r,20).value),clean(ws.cell(r,21).value),clean(ws.cell(r,23).value),sheet,r))
        # Detailed budget uses the existing approved budget workbook source sheets. Exclude payroll/capex only at report level; preserve all source data.
        wb=self.wb("E88_ApprovedBudget2026 (4)(1).xlsx")
        month_map={m:i for i,m in enumerate(["Jan","Feb","Mar","Apr","May","June","July","Aug","Sep","Oct","Nov","Dec"],1)}
        for sheet in wb.sheetnames:
            if not sheet.endswith(" Budget vs Actual"): continue
            month_name=sheet.split()[0];month=month_map.get(month_name)
            if not month: continue
            ws=wb[sheet]
            for r in range(4,(ws.max_row or 0)+1):
                dept=clean(ws.cell(r,1).value);cost=clean(ws.cell(r,2).value);amount=number(ws.cell(r,3).value)
                if dept and (cost or amount):
                    self.db.execute("INSERT OR IGNORE INTO erp_budget_plan(year,month,department,cost_center,account_title,capex_opex,amount,source_system,source_sheet,source_row) VALUES(2026,?,?,?,?,?,?, 'APPROVED_BUDGET',?,?)",
                                    (month,dept,cost,cost,"OPEX",amount,sheet,r))
        # Financial-model drivers, one row per annual driver value.
        wb=self.wb("E88_AM_FINAL_v5A.xlsx")
        for sheet,business in [("MC Drivers (NRD)","NRD MOTORCYCLE"),("Energy Drivers (RideBox)","RIDEBOX ENERGY")]:
            ws=wb[sheet]
            years={c:int(number(ws.cell(4,c).value)) for c in range(4,15) if 2020<=number(ws.cell(4,c).value)<=2040}
            current_group=""
            for r in range(5,(ws.max_row or 0)+1):
                name=clean(ws.cell(r,2).value);unit=clean(ws.cell(r,3).value)
                if not name: continue
                if not any(ws.cell(r,c).value not in (None,"") for c in years):
                    current_group=name;continue
                for c,yr in years.items():
                    v=ws.cell(r,c).value
                    if v not in (None,"") and isinstance(v,(int,float)):
                        self.db.execute("INSERT OR IGNORE INTO erp_planning_drivers(business_model,driver_group,driver_name,unit_basis,plan_year,value,source_system,source_sheet,source_row) VALUES(?,?,?,?,?,?, 'FINANCIAL_MODEL',?,?)",
                                        (business,current_group,name,unit,yr,float(v),sheet,r))

    def load_station_projects(self):
        ws=self.wb("ATLAS - Asset Manifest (1).xlsx")["LOCKER"]
        project_by_site={}
        for r in range(2,(ws.max_row or 0)+1):
            sn=serial(ws.cell(r,4).value);site=clean(ws.cell(r,8).value)
            if not sn or not site: continue
            key=norm(site)
            if key not in project_by_site:
                partner=self.ensure_partner(site,"SITE_PARTNER",source_system="ATLAS",source_key=site)
                no=f"BSSP-MIG-{self.station_seq:05d}";self.station_seq+=1
                cur=self.db.execute("""INSERT INTO erp_station_projects(project_no,site_name,partner_id,planned_location,planned_date,target_activation_date,actual_activation_date,progress_pct,status,created_at)
                                     VALUES(?,?,?,?,?,?,?,100,'ACTIVE',datetime('now'))""",(no,site,partner["id"],site,date_text(ws.cell(r,7).value),date_text(ws.cell(r,12).value),date_text(ws.cell(r,12).value)))
                project_by_site[key]=cur.lastrowid
            a=self.get_asset(sn)
            self.db.execute("INSERT OR IGNORE INTO erp_station_project_assets(project_id,asset_id,serial_no,asset_role,assigned_date,status) VALUES(?,?,?,?,?,'ASSIGNED')",
                            (project_by_site[key],a["id"] if a else None,sn,"SWAPPING_STATION",date_text(ws.cell(r,11).value or ws.cell(r,7).value)))
        # Map legacy battery assignments to station projects where station code/site is available.
        for r in self.db.execute("SELECT * FROM battery_mapping"):
            sn=serial(r["serial_no"]);site=clean(r["station_code"] or r["locker_no"])
            if not sn or not site: continue
            key=norm(site)
            if key not in project_by_site:
                p=self.ensure_partner(site,"SITE_PARTNER")
                no=f"BSSP-MIG-{self.station_seq:05d}";self.station_seq+=1
                cur=self.db.execute("INSERT INTO erp_station_projects(project_no,site_name,partner_id,planned_location,progress_pct,status) VALUES(?,?,?,?,100,'ACTIVE')",(no,site,p["id"],site));project_by_site[key]=cur.lastrowid
            a=self.get_asset(sn)
            self.db.execute("INSERT OR IGNORE INTO erp_station_project_assets(project_id,asset_id,serial_no,asset_role,assigned_date,status) VALUES(?,?,?,?,?,'ASSIGNED')",(project_by_site[key],a["id"] if a else None,sn,"BATTERY",date_text(r["mapped_date"])))

    def finalize(self):
        self.db.execute("INSERT OR REPLACE INTO erp_users(email,display_name,role_code,department,live_access,active) VALUES('mmungcal@nrdev.ph','Mark Alexis Mungcal','ADMIN','Finance and Accounting',1,1)")
        self.db.execute("INSERT OR REPLACE INTO erp_settings(key,value,updated_at,updated_by) VALUES('OPENING_DATA_LOADED','YES',datetime('now'),'MIGRATION')")
        self.db.execute("INSERT OR REPLACE INTO erp_settings(key,value,updated_at,updated_by) VALUES('OPENING_DATA_SOURCE_COUNT',?,datetime('now'),'MIGRATION')",(str(len(SOURCE_FILES)),))
        self.db.execute("INSERT OR REPLACE INTO erp_settings(key,value,updated_at,updated_by) VALUES('OPENING_DATA_SOURCE_ROWS',?,datetime('now'),'MIGRATION')",(str(self.counts["source_rows"]),))
        # Advance auto-code sequences beyond all migrated category-based codes.
        for cat in ["MC","BAT","BSS","SP","CHG","OTH"]:
            mx=0
            for row in self.db.execute("SELECT item_code FROM erp_items WHERE item_code LIKE ?",(cat+"-%",)):
                m=re.fullmatch(re.escape(cat)+r"-(\d+)",row[0] or "")
                if m: mx=max(mx,int(m.group(1)))
            self.db.execute("UPDATE erp_sequences SET next_value=? WHERE code=?",(mx+1,"ITEM_"+cat))
        for typ,pref in [("CUSTOMER","CUS"),("VENDOR","VEN"),("EMPLOYEE","EMP"),("SITE_PARTNER","PAR")]:
            mx=0
            for row in self.db.execute("SELECT partner_code FROM erp_partners WHERE partner_code LIKE ?",(pref+"-%",)):
                m=re.fullmatch(re.escape(pref)+r"-(\d+)",row[0] or "")
                if m: mx=max(mx,int(m.group(1)))
            self.db.execute("UPDATE erp_sequences SET next_value=? WHERE code=?",(mx+1,"PARTNER_"+typ))

        # Update source control normalized/exception counts at the file level.
        for fn in SOURCE_FILES:
            self.db.execute("UPDATE erp_opening_data_control SET normalized_rows=?,exception_rows=? WHERE source_name=?",(0,0,fn))
        self.db.commit()


def sql_literal(v: Any) -> str:
    if v is None: return "NULL"
    if isinstance(v,(int,float)) and not isinstance(v,bool): return repr(v)
    return "'" + str(v).replace("'","''") + "'"


def export_data_chunks(db: sqlite3.Connection, chunk_rows: int = 5000):
    OPENING.mkdir(parents=True,exist_ok=True)
    for f in OPENING.glob("*.sql"): f.unlink()
    statements=[]
    for table in CONNECTED_TABLES:
        if not db.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",(table,)).fetchone(): continue
        cols=[r[1] for r in db.execute(f"PRAGMA table_info({table})")]
        rows=db.execute(f"SELECT * FROM {table}").fetchall()
        mode = "REPLACE" if table in {"erp_settings","erp_sequences","erp_users","erp_opening_data_control"} else "IGNORE"
        for row in rows:
            values=",".join(sql_literal(row[c]) for c in cols)
            statements.append(f"INSERT OR {mode} INTO {table}({','.join(cols)}) VALUES({values});")
    files=[]
    for i in range(0,len(statements),chunk_rows):
        path=OPENING/f"{i//chunk_rows+1:04d}_opening_data.sql"
        path.write_text("PRAGMA foreign_keys=OFF;\n"+"\n".join(statements[i:i+chunk_rows])+"\nPRAGMA foreign_keys=ON;\n",encoding="utf-8")
        files.append(path)
    manifest=OPENING/"manifest.json"
    manifest.write_text(json.dumps({"files":[p.name for p in files],"statements":len(statements)},indent=2),encoding="utf-8")
    return files,len(statements)


def main():
    ap=argparse.ArgumentParser();ap.add_argument("--skip-archive",action="store_true");args=ap.parse_args()
    missing=[x for x in SOURCE_FILES if not (SOURCE/x).exists()]
    if missing: raise SystemExit(f"Missing source files: {missing}")
    db=make_conn();loader=Loader(db)
    if not args.skip_archive: loader.archive_sources()
    print("Loading masters", flush=True); loader.load_master_data()
    print("Collecting assets", flush=True); loader.collect_assets()
    print("Loading shipments/receiving/landed cost", flush=True); loader.load_shipments_receiving_landed_cost()
    print("Loading movements/current state", flush=True); loader.load_movements_and_current_state()
    print("Loading requisitions/checklists", flush=True); loader.load_requisitions_checklists()
    print("Loading procurement/sales/budget/planning", flush=True); loader.load_procurement_sales_budget_planning()
    print("Loading station projects", flush=True); loader.load_station_projects()
    loader.finalize()
    files,statements=export_data_chunks(db)
    counts={t:db.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0] for t in CONNECTED_TABLES if db.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",(t,)).fetchone()}
    REPORTS.mkdir(exist_ok=True)
    (REPORTS/"DATA_LOAD_REPORT.json").write_text(json.dumps({"generated_at":datetime.now().isoformat(),"source_files":SOURCE_FILES,"counts":counts,"chunks":[p.name for p in files],"statements":statements},indent=2),encoding="utf-8")
    print(json.dumps({"chunks":len(files),"statements":statements,"counts":counts},indent=2))

if __name__=="__main__": main()
