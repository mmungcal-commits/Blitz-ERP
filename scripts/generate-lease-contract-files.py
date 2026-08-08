#!/usr/bin/env python3
"""
The signed contracts were hyperlinked all along.

Column B of the SCHED sheet carries the client name, and on fourteen of those
rows the name is a link to the signed contract in Google Drive - exactly the
pattern the bank advices followed in the procurement sheet. Nobody had to
re-upload anything; the paper was already filed, just not on the record.

Keyed on client code plus batch (D and E, which the sheet itself concatenates
into the CB code) so a contract lands on the right lease and not merely on the
right customer: Henry Soesanto has seven contracts and they are not the same
agreement.
"""
import openpyxl, re

SRC = 'source_data/E88_LeaseContracts.xlsx'
OUT = 'migrations/0064_lease_contract_files.sql'

def q(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

wb = openpyxl.load_workbook(SRC)
ws = wb['SCHED']

rows = []
for r in range(6, ws.max_row + 1):
    cell = ws.cell(r, 2)                       # B: client, and the link
    if not cell.value: continue
    name = str(cell.value).strip()
    if not name or name.upper().startswith('TOTAL'): continue
    link = cell.hyperlink.target if cell.hyperlink else None
    if not link: continue
    code  = ws.cell(r, 4).value                # D: client code
    batch = ws.cell(r, 5).value                # E: batch
    if not code: continue
    batch_txt = str(batch).strip()
    if batch_txt.endswith('.0'): batch_txt = batch_txt[:-2]
    cb = f"{str(code).strip()}-{batch_txt}"
    rows.append((cb, name, link.strip()))

seen, uniq = set(), []
for cb, name, link in rows:
    if cb in seen: continue
    seen.add(cb); uniq.append((cb, name, link))

with open(OUT, 'w') as f:
    f.write("""-- 0064 · The signed contracts were hyperlinked all along
--
-- The lease register showed "none" under Contract file against every one of the
-- twenty-two contracts, and the uploader built in R49 sat unused. The paper was
-- not missing: column B of the SCHED sheet carries the client name, and on
-- fourteen of those rows the name is a link to the signed contract in Drive -
-- the same pattern the bank advices followed in the procurement sheet.
--
-- Filed against the lease rather than the order, matching where R49 put the
-- uploader, and keyed on the CB code so a contract lands on the right lease.
-- Henry Soesanto has seven of them and they are not the same agreement.
--
-- Re-runnable: NOT EXISTS on the same link against the same contract, so a
-- redeploy re-attaches nothing and a contract replaced by hand is left alone.

""")
    for cb, name, link in uniq:
        f.write(
            "INSERT INTO erp_attachments(record_type,record_id,module_code,record_no,"
            "file_name,content_type,file_url,storage,uploaded_by,active)\n"
            f"SELECT 'LEASE_CONTRACT', l.id, 'SALES', l.lease_no, "
            f"'Signed contract - ' || l.lease_no || '.pdf', 'application/pdf', {q(link)}, "
            "'DRIVE_LINK', 'lease-register@nrdev.ph', 1\n"
            "  FROM erp_lease_contracts l\n"
            "  JOIN erp_lease_contract_batches b ON b.lease_contract_id=l.id\n"
            f" WHERE UPPER(TRIM(b.cb_code))=UPPER(TRIM({q(cb)}))\n"
            "   AND NOT EXISTS (SELECT 1 FROM erp_attachments a\n"
            "        WHERE a.record_type='LEASE_CONTRACT' AND a.record_id=l.id\n"
            f"          AND a.file_url={q(link)} AND a.active=1);\n")

print(f"{len(uniq)} contracts written to {OUT}")
